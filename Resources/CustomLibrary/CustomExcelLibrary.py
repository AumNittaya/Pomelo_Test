import openpyxl

class CustomExcelLibrary:

    # Open Excel Workbook
    def open_excel_workbook(self, filePath):
        return openpyxl.load_workbook(filePath, data_only=True)

    # Select worksheet from workbook
    def select_excel_worksheet(self, wb, sheetName):
        return wb[sheetName]

    # Save excel workbook
    def save_excel_workbook(self, wb, filePath):
        wb.save(filePath)

    # Close excel workbook
    def close_excel_workbook(self, wb):
        wb.close()

    # Write excel worksheet by Cell index (Row and column is require)
    def write_cell(self, ws, row, column, data):
        ws.cell(int(row), int(column)).value = data

    # Write excel worksheet by cell ref (cell ref is require)
    def write_cell_by_cell_detail(self, ws, cellRef, data):
        ws.cell(int(cellRef['row']), int(cellRef['column'])).value = data
    
    def read_excel_sheet_data_from_file(self, file, sheetName, cellRef=False, testcaseId=None, headerRow=1, testcaseColumn=1):
        wb = self.open_excel_workbook(file)
        ws = self.select_excel_worksheet(wb, sheetName)
        return self.read_excel_sheet(ws, cellRef, testcaseId, headerRow, testcaseColumn)

    def read_excel_sheet_data_from_workbook(self, wb, sheetName, cellRef=False, testcaseId=None, headerRow=1, testcaseColumn=1):
        ws = self.select_excel_worksheet(wb, sheetName)
        return self.read_excel_sheet(ws, testcaseId, headerRow, cellRef, testcaseColumn)

    def read_excel_sheet(self, ws, cellRef=False, testcaseId=None, headerRow=1, testcaseColumn=1):
        dataArray = []
        headerArray = []
        for i in range(ws.max_column):
            headerCell = ws.cell(headerRow, i+1).value
            if headerCell != None:
                headerCell = headerCell.replace(" ", "_")
                headerArray.append(headerCell)
        for i in range(ws.max_row):
            rowIndex = i + 1
            if rowIndex <= headerRow:
                continue
            dataDict = {}
            testcase = ws.cell(rowIndex, testcaseColumn).value
            if testcaseId != None:
                if testcase != testcaseId:
                    continue
            for i in range(len(headerArray)):
                dataCellIndex = i + 1
                headerVal = headerArray[i]
                dataCellVal = ws.cell(rowIndex, dataCellIndex).value
                if dataCellVal == '' or dataCellVal == ' ' or dataCellVal == '/':
                    dataCellVal = '/'
                if cellRef == True:
                    cellDict = {}
                    cellDict["row"] = str(rowIndex)
                    cellDict["column"] = str(dataCellIndex)
                    cellDict["value"] = str(dataCellVal)
                    cellDict["columnName"] = headerVal
                    dataDict[headerVal] = cellDict
                else:
                    dataDict[headerVal] = str(dataCellVal)
            dataDict['Row_Index'] = rowIndex
            dataArray.append(dataDict)
        return dataArray

    def find_column_index_in_row(self, ws, rowIndex, columnName):
        maxCol = ws.max_column
        columnNameSearch = columnName.replace(" ", "_")
        columnNameSearch = columnNameSearch.lower()
        for i in range(maxCol):
            index = i + 1
            columnNameRes = ws.cell(int(rowIndex), index).value
            columnNameRes = columnNameRes.replace(" ", "_")
            columnNameRes = columnNameRes.lower()
            if columnNameRes == columnNameSearch:
                return index
        return "Not found data in column"

    def find_row_index_in_column(self, ws, columnIndex, uniqueKeyVal):
        maxRow = ws.max_row
        for i in range(maxRow):
            index = i + 1
            if ws.cell(index,columnIndex).value == uniqueKeyVal:
                return index
        return "Not found data in row"

    def get_last_row(self, ws):
        return ws.max_row

    def get_last_column(self, ws):
        return ws.max_column

    def get_cell_value(self, ws, row, column):
        return ws.cell(row, column).value

    def filter_json_array(self, json_data, searchKey, searchValue):
        output_json = list(filter(lambda x: x[searchKey] in str(searchValue), json_data))
        return output_json
        
    def filter_json_with_cell_ref(self, json_data, searchKey, searchValue):
        output_json = list(filter(lambda x: x[searchKey]['value'] in str(searchValue), json_data))
        return output_json